import fitz
from datetime import datetime
import openpyxl
import os
import sys

#retorna un string con la fecha/hora del sistema, 
# se usa para evitar duplicados en archivos
def now():
    fecha=datetime.now()
    now=fecha.strftime('%d-%m-%Y-%H-%M-%S')
    return now

#Verifica si una carpeta ya está creada, si no es así la crea
def checkCreateFolder(nameFolder):
  try: os.mkdir(nameFolder)
  except: pass
  location=f'./{nameFolder}'
  return location

#De una lista de palabras a buscar en un PDF retorna una lista de coords
def ccWords(listWords,pdfAdress):

    file=fitz.open(pdfAdress)
    coord=[]

    for i, page in enumerate(file.pages(), start=1):
        coord.append(f'page: {i}')
        pageword=page.get_text('words')
        for word in pageword:
            if word[4] in listWords:
                coord.append(word)

    return coord

#Guarda elementos de una lista en un txt
def saveListTxt(lista):
    ahora=now()
    with open(f"./file_{ahora}.txt", "w") as output:
        for row in lista:
            output.write(str(row)+'\n')

#Devuelve un string a partir de una frase y dos strings de referencia que
#delimitan el inicio y el fin del string. No toma los string de incio y fin.
#Nota: el código es mejorable.
def wordBetween(pStart,pEnd,phrase,include):
    char1=pStart[0]
    char2=pEnd[0]
    val1,val2=False,False
    finalWord=''
    
    for i, char in enumerate(phrase):
        if char1==char and val1==False:
            for j, letter in enumerate(pStart):
                if letter==phrase[i+j]: val1=True
                else: 
                    val1=False
                    break

        if val1: finalWord+=char

        if char2==char and val2==False and val1==True:
            for j, letter in enumerate(pEnd):
                if letter==phrase[i+j]: val2=True
                else: 
                    val2=False
                    break
        
        if val2: break

    finalWord=finalWord[len(pStart):]

    if include==True: finalWord=pStart+finalWord+pEnd

    return finalWord

#A partir de una lista de coordenadas XY y la ubicación de un archivo PDF busca
#todas las palabras que estén en esas coordenadas hoja por hoja y almacena todas
#esas palabras en un diccionario con la metadata de cada palabra y retorna el
#diccionario. 
def wordPDFdic (listCC,pdfAddress):
    file=fitz.open(pdfAddress)
    wordDic={}

    for i, page in enumerate(file.pages(), start=1):
        wordList=[]
        xy=[1,2]
        for word in page.get_text('words'):
            xy[0]=word[0]
            xy[1]=word[1]
            if xy in listCC: wordList.append(word[4])
        wordDic[i]=wordList

    return wordDic

#Crea un diccionario con toda su meta data a partir de una coordenada xy
def listWordCC(x,y,pdfAddress):
    file=fitz.open(pdfAddress)
    wordList={}

    for nPage, page in enumerate(file,start=1):
        pageWords=page.get_text('words')
        for word in pageWords:
            if word[0]==x and word[1]==y:
                wordList[nPage]=word

    return wordList

#A partir de un diccionario de campos guarda una tabla en un excel
def saveExceltoDic(dicWord,filename,address):
    wb=openpyxl.Workbook()
    ws=wb.create_sheet('BD')
    del wb['Sheet']

    for i, campo in enumerate(dicWord, start=1):
        ws.cell(row=1,column=i).value=campo
        for _, items in enumerate(dicWord[campo],start=0):
            word=dicWord[campo][4]
            ws.cell(row=items+1,column=i).value=word
    
    wb.save(f'{address}/{filename}.xlsx')

#funcion exclusiva para trabajar con el guardado en la la interface.
def saveExceltoDicInterface(dicWord,address):
    wb=openpyxl.Workbook()
    ws=wb.create_sheet('BD')
    del wb['Sheet']

    for i, campo in enumerate(dicWord, start=1):
        ws.cell(row=1,column=i).value=campo
        for _, items in enumerate(dicWord[campo],start=0):
            word=dicWord[campo][items][4]
            ws.cell(row=items+1,column=i).value=word
    
    wb.save(address)

#Extrae las imagenes de un pdf y las guarda en una carpeta dada como input.
#Tiene la opción de filtrar las imagenes inferiores a x dimensiones dadas como [x,y].
#También permite asignarles nombre a partir de una lista de nombres, en caso que
#se pase una lista vacía se le asignará un nombre genérico a la imagen.
def extractImgPDF(addressPDF,destination,dimention,listName):
  
  file=fitz.open(addressPDF)
  number_of_pages = len(file)
  j=0

  #iterating through each page in the pdf
  for number in range(number_of_pages):
    
    #muestra el % de progreso
    n=(number/number_of_pages)*100
    n=round(n,2)
    print('En proceso: ', end=' ')
    sys.stdout.write('{}%\r'.format(n))
    if number==number_of_pages-1: print('En proceso: 100%    ')

    #iterating through each image in every page of PDF       
    for idimg,img in enumerate(file.get_page_images(number)):

        #filtro por dimensiones de imagen
        if img[2] <dimention[0] or img[3]<dimention[1]: continue
        
        try: name=listName[j]
        except: name=f'page_{number+1}-img-{idimg}'

        xref = img[0]
        image = fitz.Pixmap(file, xref)

        #if it is a is GRAY or RGB image
        if image.n < 5:        
            image.save(f"{destination}/{name}.png")
            
        #if it is CMYK: convert to RGB first
        else:
            new_image = fitz.Pixmap(fitz.csRGB, image)
            new_image.save(f"{destination}/{name}.png")

        j+=1


        
    

